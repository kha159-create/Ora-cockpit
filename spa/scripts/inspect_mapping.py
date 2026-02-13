import pandas as pd
import sys

file_path = 'C:/Users/Orange1/Downloads/mapping.xlsx'
output_path = 'C:/Users/Orange1/Desktop/Ora-cockpit/spa/scripts/inspect_output.txt'

with open(output_path, 'w', encoding='utf-8') as f:
    f.write(f"Reading {file_path}...\n")
    try:
        df = pd.read_excel(file_path, nrows=5)
        f.write(f"Columns: {df.columns.tolist()}\n")
        for i, row in df.iterrows():
            f.write(f"Row {i}: {row.tolist()}\n")
    except Exception as e:
        f.write(f"Pandas failed: {e}\n")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            f.write("Using openpyxl...\n")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 5:
                    f.write(f"Row {i}: {row}\n")
                else:
                    break
        except Exception as e2:
            f.write(f"Openpyxl failed: {e2}\n")
